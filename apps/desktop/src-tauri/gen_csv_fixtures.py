#!/usr/bin/env python3
"""Generates real fixture files for MiniOffice integration tests."""
import csv, io, os

OUT = os.path.join(os.path.dirname(__file__), "fixtures")
os.makedirs(OUT, exist_ok=True)

def w(name: str, data: bytes):
    with open(os.path.join(OUT, name), "wb") as f:
        f.write(data)
    print("wrote", name, len(data), "bytes")

# 1. basic.csv — plain comma file
w("basic.csv", b"name,age,city\nJohn,32,Delhi\nSarah,28,Mumbai\nMike,41,Pune\n")

# 2. nasty.csv — RFC4180 stress: quoted fields, embedded commas/quotes/newlines,
#    leading/trailing spaces preserved by quoting, unicode
nasty_rows = [
    ["name", "notes", "score"],
    ["Smith, John", 'He said "hello" twice', "91"],
    ["Ana María", "multi-line\nsecond line, with comma", "87.5"],
    ["日本語 テスト", "trailing spaces kept  ", ""],
    ["emoji ✓ row", "quote inside \"quoted\" text", "-12"],
]
buf = io.StringIO()
csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\n").writerows(nasty_rows)
w("nasty.csv", buf.getvalue().encode("utf-8"))

# 3. bom.csv — UTF-8 BOM + semicolons would be wrong; keep commas but BOM
rows = [["product", "price"], ["Kaffee ☕", "3,20"], ["Tee", "2,10"]]
body = csv.writer(io.StringIO(), lineterminator="\n")
s = io.StringIO()
csv.writer(s, lineterminator="\n").writerows(rows)
w("bom.csv", b"\xef\xbb\xbf" + s.getvalue().encode("utf-8"))

# 4. semi.csv — semicolon delimiter (European Excel export style)
w("semi.csv", "name;stadt;umsatz\nMüller;München;1200,50\nSchmidt;Berlin;980,00\n".encode("utf-8"))

# 5. onecol.csv — single column, no delimiter characters at all (regression:
#    the sniffer used to reject these)
w("onecol.csv", "alpha\nbeta\ngamma\n".encode())

# 6. crlf.csv — Windows line endings
w("crlf.csv", b"a,b\r\n1,2\r\n3,4\r\n")

# 7. empty.csv — zero bytes
w("empty.csv", b"")

# ---- XLSX with formulas but NO cached values (openpyxl never computes) ----
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    rows = [
        ["item", "region", "qty", "unit_price"],
        ["widget", "north", 10, 2.5],
        ["widget", "south", 6, 2.5],
        ["gizmo", "north", 4, 7.25],
        ["gadget", "east", 15, 1.1],
        ["widget", "east", 8, 2.5],
    ]
    for r in rows:
        ws.append(r)
    # formulas without cached values (openpyxl stores only the formula string)
    ws["F1"] = "total_qty"
    ws["F2"] = "=SUM(C2:C6)"
    ws["G1"] = "high_value"
    ws["G2"] = '=IF(D3*10>50,"High","Low")'
    ws["H1"] = "lookup"
    ws["H2"] = '=VLOOKUP("gizmo",A2:D6,4,FALSE)'
    wb.create_sheet("Summary")
    ws2 = wb["Summary"]
    ws2["A1"] = "widgets_north"
    ws2["B1"] = '=SUMIF(Sales!B2:B6,"north",Sales!C2:C6)'
    ws2["A2"] = "index_match"
    ws2["B2"] = '=INDEX(Sales!D2:D6,MATCH("gadget",Sales!A2:A6,0))'
    wb.save(os.path.join(OUT, "formulas.xlsx"))
    print("wrote formulas.xlsx")
except ImportError:
    print("openpyxl missing; skipping formulas.xlsx")

print("done")
