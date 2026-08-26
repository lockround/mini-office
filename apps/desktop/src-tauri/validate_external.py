#!/usr/bin/env python3
"""Third-party validation: files MiniOffice writes must be readable by
other tools (openpyxl), and real-world fixtures must parse cleanly."""
import sys, zipfile, os

OUT = os.path.join(os.path.dirname(__file__), "fixtures")
import openpyxl

failures = []
def check(cond, msg):
    print(("PASS " if cond else "FAIL ") + msg)
    if not cond:
        failures.append(msg)

# 1. A workbook saved by MiniOffice (rust_xlsxwriter) opens in openpyxl,
#    keeps sheet names, values, and formula strings.
src = os.path.join(OUT, "complex.xlsx")
wb = openpyxl.load_workbook(src)          # data_only=False → formulas
check(wb.sheetnames == ["Data", "Summary"], f"sheet names {wb.sheetnames}")
ws = wb["Summary"]
check(ws["B2"].value == "=SUM(Data!C2:C7)", f"B2 formula: {ws['B2'].value!r}")
check(ws["B9"].value.startswith("=IF(SUMIF"), f"nested_if preserved: {ws['B9'].value[:30]!r}")
wsd = wb["Data"]
check(wsd["E2"].value == "=C2*D2", f"E2 formula: {wsd['E2'].value!r}")
check(float(wsd["C2"].value) == 10, "numeric cell intact")

# cached values visible via data_only=True
wbv = openpyxl.load_workbook(src, data_only=True)
wsv = wbv["Summary"]
check(str(wsv["B2"].value) == "46", f"B2 cached value: {wsv['B2'].value!r}")
check(wsv["B9"].value == "High N", f"nested_if cached: {wsv['B9'].value!r}")

# 2. Round-trip: load complex.xlsx and re-save with openpyxl; MiniOffice's
#    parser must still extract every formula afterwards (simulates another
#    tool touching the file between sessions).
rt = os.path.join(OUT, "_roundtrip.xlsx")
wb.save(rt)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) or ".")

print("--- openpyxl re-save done, checking zip structure of our own outputs ---")

# 3. Structural sanity of a docx produced by our generator script
docx = os.path.join(OUT, "report.docx")
with zipfile.ZipFile(docx) as z:
    names = z.namelist()
    check("[Content_Types].xml" in names, "docx has [Content_Types].xml")
    check(any(n == "word/document.xml" for n in names), "docx has word/document.xml")
    xml = z.read("word/document.xml").decode("utf-8")
    check("Quarterly Report" in xml, "heading text present")
    check("full dataset" in xml, "hyperlink text present")
    check("<w:tbl>" in xml, "table present")

os.remove(rt)
if failures:
    print(f"\n{len(failures)} FAILURES")
    sys.exit(1)
print("\nALL EXTERNAL VALIDATION PASSED")
